import sys
print("Python version:", sys.version)

from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for
from models import Database
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import io

app = Flask(__name__)
app.secret_key = 'ecole_mont_sion_secret_key'

db = Database()

MATIERES = [
    'Mathématiques',
    'Communication écrite',
    'Lecture',
    'Anglais',
    'SVT',
    'Histoire-géographie',
    'Espagnol',
    'EPS',
    'Conduite'
]

# ---------- ACCUEIL ----------
@app.route('/')
def accueil():
    return render_template('accueil.html')

# ---------- INSCRIPTION ----------
@app.route('/inscription')
def inscription():
    return render_template('inscription.html')

@app.route('/inscrire_ecolier', methods=['POST'])
def inscrire_ecolier():
    data = request.json
    ecolier = {
        'nom': data['nom'],
        'prenoms': data['prenoms'],
        'sexe': data['sexe'],
        'date_naissance': data['date_naissance'],
        'classe': data['classe'],
        'numero_parents': data['numero_parents'],
        'montant_scolarite': data['montant_scolarite'],
        'nom_enregistreur': data['nom_enregistreur']
    }
    db.add_ecolier(ecolier)
    return jsonify({'success': True})

@app.route('/inscrire_eleve', methods=['POST'])
def inscrire_eleve():
    data = request.json
    eleve = {
        'nom': data['nom'],
        'prenoms': data['prenoms'],
        'sexe': data['sexe'],
        'date_naissance': data['date_naissance'],
        'classe': data['classe'],
        'numero_parents': data['numero_parents'],
        'montant_scolarite': data['montant_scolarite'],
        'nom_enregistreur': data['nom_enregistreur']
    }
    db.add_eleve(eleve)
    return jsonify({'success': True})

# ---------- LISTES ----------
@app.route('/liste_eleves')
def liste_eleves():
    eleves = db.get_eleves()
    return render_template('liste_eleves.html', eleves=eleves)

@app.route('/liste_ecoliers')
def liste_ecoliers():
    ecoliers = db.get_ecoliers()
    return render_template('liste_ecoliers.html', ecoliers=ecoliers)

# ---------- SCOLARITÉ (CORRIGÉ) ----------
@app.route('/scolarite')
def scolarite():
    students = db.get_all()
    for s in students:
        try:
            montant = int(str(s.get('montant_scolarite', '0')).strip())
        except ValueError:
            montant = 0
        total = db.get_total_paid(s)
        s['total_paid'] = total
        s['reste'] = montant - total
    return render_template('scolarite.html', students=students)

# ---------- NOTES (CORRIGÉ) ----------
@app.route('/notes')
def notes():
    ecoliers = db.get_ecoliers()
    eleves = db.get_eleves()
    classes_ecoliers = sorted(set([e['classe'] for e in ecoliers]))
    classes_eleves = sorted(set([e['classe'] for e in eleves]))
    return render_template('notes.html', classes_ecoliers=classes_ecoliers, classes_eleves=classes_eleves, matieres=MATIERES)

@app.route('/get_students_by_class', methods=['POST'])
def get_students_by_class():
    data = request.json
    classe = data['classe']
    is_ecolier = data['is_ecolier']
    students = [s for s in (db.get_ecoliers() if is_ecolier else db.get_eleves()) if s['classe'] == classe]
    return jsonify({'students': students})

@app.route('/save_notes', methods=['POST'])
def save_notes():
    data = request.json
    for note in data['notes']:
        db.add_note(note['student_id'], note['student_type'], note['classe'], note['matiere'], note['note'])
    return jsonify({'success': True})

# ---------- VUE NOTES (CORRIGÉ) ----------
@app.route('/vue_notes')
def vue_notes():
    notes = db.get_notes()
    classes = sorted(set([n['classe'] for n in notes]))
    matieres = sorted(set([n['matiere'] for n in notes]))
    return render_template('vue_notes.html', classes=classes, matieres=matieres)

@app.route('/get_notes_by_class', methods=['POST'])
def get_notes_by_class():
    data = request.json
    classe = data['classe']
    matiere = data['matiere']
    all_students = db.get_ecoliers() if classe in ['maternelle','CI','CP','CE1','CE2','CM1','CM2'] else db.get_eleves()
    students = []
    for s in all_students:
        if s['classe'] == classe:
            notes = db.get_student_notes(s['id'], 'ecolier' if classe in ['maternelle','CI','CP','CE1','CE2','CM1','CM2'] else 'eleve')
            note_val = next((n['note'] for n in notes if n['matiere'] == matiere), None)
            students.append({'id': s['id'], 'nom': s['nom'], 'prenoms': s['prenoms'], 'note': note_val})
    return jsonify({'students': students})

# ---------- SAUVEGARDE ----------
@app.route('/sauvegarde')
def sauvegarde():
    data = db.load_data()
    stats = {'ecoliers': len(data['ecoliers']), 'eleves': len(data['eleves']), 'notes': len(data['notes'])}
    return render_template('sauvegarde.html', stats=stats)

# ---------- IMPORT / EXPORT ----------
@app.route('/import_excel', methods=['GET', 'POST'])
def import_excel():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            data = {'ecoliers': [], 'eleves': [], 'notes': []}
            db.save_data(data)

            if 'Écoliers' in wb.sheetnames:
                for row in wb['Écoliers'].iter_rows(min_row=2, values_only=True):
                    if row[1] and row[2]:
                        db.add_ecolier({
                            'nom': row[1], 'prenoms': row[2], 'sexe': row[3], 'date_naissance': row[4],
                            'classe': row[5], 'numero_parents': str(row[6]), 'montant_scolarite': str(row[7]),
                            'nom_enregistreur': 'Import Excel'
                        })

            if 'Élèves' in wb.sheetnames:
                for row in wb['Élèves'].iter_rows(min_row=2, values_only=True):
                    if row[1] and row[2]:
                        db.add_eleve({
                            'nom': row[1], 'prenoms': row[2], 'sexe': row[3], 'date_naissance': row[4],
                            'classe': row[5], 'numero_parents': str(row[6]), 'montant_scolarite': str(row[7]),
                            'nom_enregistreur': 'Import Excel'
                        })

            if 'Notes' in wb.sheetnames:
                for row in wb['Notes'].iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[2] and row[3]:
                        all_students = db.get_all()
                        for s in all_students:
                            if f"{s['nom']} {s['prenoms']}" == row[0]:
                                db.add_note(s['id'], s['type'], row[1], row[2], str(row[3]))
                                break

            flash('✅ Importation réussie avec succès !', 'success')
            return redirect(url_for('sauvegarde'))
    return render_template('import_excel.html')

@app.route('/export_excel')
def export_excel():
    wb = openpyxl.Workbook()
    # Écoliers
    ws = wb.active
    ws.title = "Écoliers"
    headers = ['ID', 'Nom', 'Prénoms', 'Sexe', 'Date de naissance', 'Classe', 'Numéro parents', 'Montant scolarité', 'Total payé', 'Reste', 'Date inscription']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for row, e in enumerate(db.get_ecoliers(), 2):
        total = db.get_total_paid(e)
        ws.cell(row=row, column=1).value = e['id']
        ws.cell(row=row, column=2).value = e['nom']
        ws.cell(row=row, column=3).value = e['prenoms']
        ws.cell(row=row, column=4).value = e['sexe']
        ws.cell(row=row, column=5).value = e['date_naissance']
        ws.cell(row=row, column=6).value = e['classe']
        ws.cell(row=row, column=7).value = e['numero_parents']
        ws.cell(row=row, column=8).value = e['montant_scolarite']
        ws.cell(row=row, column=9).value = total
        ws.cell(row=row, column=10).value = int(e['montant_scolarite']) - total
        ws.cell(row=row, column=11).value = e['date_inscription']

    # Élèves
    ws = wb.create_sheet("Élèves")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for row, e in enumerate(db.get_eleves(), 2):
        total = db.get_total_paid(e)
        ws.cell(row=row, column=1).value = e['id']
        ws.cell(row=row, column=2).value = e['nom']
        ws.cell(row=row, column=3).value = e['prenoms']
        ws.cell(row=row, column=4).value = e['sexe']
        ws.cell(row=row, column=5).value = e['date_naissance']
        ws.cell(row=row, column=6).value = e['classe']
        ws.cell(row=row, column=7).value = e['numero_parents']
        ws.cell(row=row, column=8).value = e['montant_scolarite']
        ws.cell(row=row, column=9).value = total
        ws.cell(row=row, column=10).value = int(e['montant_scolarite']) - total
        ws.cell(row=row, column=11).value = e['date_inscription']

    # Notes
    ws = wb.create_sheet("Notes")
    note_headers = ['Étudiant', 'Classe', 'Matière', 'Note', 'Date']
    for col, header in enumerate(note_headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    all_notes = db.get_notes()
    name_map = {s['id']: f"{s['nom']} {s['prenoms']}" for s in db.get_all()}
    for row, n in enumerate(all_notes, 2):
        ws.cell(row=row, column=1).value = name_map.get(n['student_id'], 'Inconnu')
        ws.cell(row=row, column=2).value = n['classe']
        ws.cell(row=row, column=3).value = n['matiere']
        ws.cell(row=row, column=4).value = n['note']
        ws.cell(row=row, column=5).value = n['date']

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    flash('✅ Exportation réussie avec succès !', 'success')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'ecole_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

# ---------- LANCEMENT ----------
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)
    
